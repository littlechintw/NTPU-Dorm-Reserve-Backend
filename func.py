#!/usr/bin/python
# -*- coding: utf-8 -*-

import sqlite3
import requests
import jwt
import os
# from dotenv import load_dotenv
import time
import base64
import json
import csv

# load_dotenv(encoding = "utf-8")
JWT_EXP_TIME = 60 * 60 # 1 hour
JWT_EXP_TIME_ONE_DAY = 60 * 60 * 12

RESERVE_START = '2022-08-26 13:00:00'
RESERVE_END = '2022-09-04 17:00:00'

def init_sqlite():
    print('Initializing sqlite3 database...')
    # Check if 'main.db' file in folder
    firstTimeLoad = False
    if not os.path.isfile('main.db'):
        firstTimeLoad = True
    print("First Time Load: {}".format(firstTimeLoad))
    conn = sqlite3.connect('main.db')
    c = conn.cursor()
    c.execute('''
                CREATE TABLE IF NOT EXISTS config (
                title   TEXT,
                data    TEXT
                )
            ''')
    
    # Check if exist and set RESERVE_START and RESERVE_END
    c.execute('''
                SELECT * FROM config
                WHERE title = ?
            ''', ('RESERVE_START',))
    if c.fetchone() == None:
        c.execute('''
                    INSERT INTO config (title, data)
                    VALUES (?, ?)
                ''', ('RESERVE_START', RESERVE_START))
    c.execute('''
                SELECT * FROM config
                WHERE title = ?
            ''', ('RESERVE_END',))
    if c.fetchone() == None:
        c.execute('''
                    INSERT INTO config (title, data)
                    VALUES (?, ?)
                ''', ('RESERVE_END', RESERVE_END))

    c.execute('''
                CREATE TABLE IF NOT EXISTS accounts (
                id            TEXT,
                passwd        TEXT,
                created       TEXT,
                status        TEXT,
                session       TEXT,
                session_time  TEXT,
                details       TEXT 
                )
            ''')
    if firstTimeLoad:
        c.execute('''
                    INSERT INTO accounts (id, passwd, created, status, session, session_time, details)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', ('first', 'time', get_nowtime_taipei_time(), 'normal', '', '', ''))
    c.execute('''
                CREATE TABLE IF NOT EXISTS member (
                id        TEXT,
                name      TEXT,
                dorm      TEXT,
                room      TEXT,
                bed       TEXT,
                status    TEXT,
                health    TEXT,
                details   JSON
                )
            ''')
    if firstTimeLoad:
        c.execute('''
                    INSERT INTO member (id, name, dorm, room, bed, status, health, details)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', ('first', 'FirstTime', 'admin', '', '', 'normal', '', ''))
    c.execute('''
                CREATE TABLE IF NOT EXISTS events (
                event_id   TEXT,
                event_name TEXT,
                date       TEXT,
                startTime  TEXT,
                endTime    TEXT,
                sun        INTEGER,
                moon       INTEGER,
                star       INTEGER,
                morn       INTEGER
                )
            ''')
    c.execute('''
                CREATE TABLE IF NOT EXISTS reserve (
                timestamp   TEXT,
                id          TEXT,
                event_id    TEXT,
                parking     TEXT
                )
            ''')
    c.execute('''
                CREATE TABLE IF NOT EXISTS logs (
                timestamp   TEXT,
                ip          TEXT,
                url         TEXT,
                id          TEXT,
                request     TEXT,
                response    TEXT
                )
            ''')
    c.execute('''
                CREATE TABLE IF NOT EXISTS loginDevice (
                user        TEXT,
                id          TEXT,
                ip          TEXT,
                iat         TEXT
                )
            ''')
    c.execute('''
                CREATE TABLE IF NOT EXISTS checkIn (
                timestamp     TEXT,
                user          TEXT,
                parking       TEXT,
                bill          TEXT,
                card          TEXT,
                visitor_id    TEXT,
                visitor_phone TEXT,
                visitor_start TEXT,
                visitor_end   TEXT
                )
            ''')
    conn.commit()
    conn.close()
def get_nowtime_taipei_time():
    return time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(time.time()))

def get_config(title):
    conn = sqlite3.connect('main.db')
    c = conn.cursor()
    c.execute('''
                SELECT data FROM config
                WHERE title = ?
            ''', (title,))
    data = c.fetchone()
    conn.close()
    return data[0]

def check_reserve_time():
    now = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
    DB_RESERVE_START = get_config('RESERVE_START')
    DB_RESERVE_END = get_config('RESERVE_END')
    if now < DB_RESERVE_START or now > DB_RESERVE_END:
        return False
    return True

# Create a short id with length
def create_id(length):
    import random
    import string
    return ''.join(random.choice(string.ascii_letters + string.digits) for _ in range(length))

def write_log(ip, url, id, request, response):
    conn = sqlite3.connect('main.db')
    c = conn.cursor()
    c.execute('''
                INSERT INTO logs (timestamp, ip, url, id, request, response)
                VALUES (datetime('now'), ?, ?, ?, ?, ?)
            ''', (ip, url, id, request, response))
    conn.commit()
    conn.close()

def verify_account_from_stu_sys(stu_id, pwd):
    logs_blue("Get date with [ " + stu_id + " ]\n")
    data = {
        'stud_num': stu_id,
        'passwd': pwd[:16],
        'x': '122',
        'y': '10'
    }

    r = requests.post(
        "https://ohs02.ntpu.edu.tw/pls/pm/stud_system.login", data=data)
    r.encoding = 'big5'

    if r.status_code != requests.codes.ok:
        logs_red('Can\'t connect to ohs02.ntpu.edu.tw!')
        return {
            "msg": "oAuth time out or get some error!",
            "err": 4001
        }

    if r.text.find('您輸入的學號或密碼有誤') > 0:
        logs_red('The ID or passwd is wrong')
        return {
            "msg": "Login failed!",
            "err": 4002
        }

    if not (r.cookies.get('STD3') and r.cookies.get('stud') and r.cookies.get('tlc')):
        logs_red('Can\'t get the cookies')
        return {
            "msg": "Cookies error!",
            "err": 4003
        }

    return {
        "msg": "OK",
        "err": 200
    }

def generate_jwt(id, device_ip):
    iat = int(time.time())
    exp = iat + JWT_EXP_TIME
    random_id = create_id(32)
    payload = {
        "account": id,
        "iat": iat,
        "exp": exp,
        "id": random_id,
    }

    # Write device info into device database
    conn = sqlite3.connect('main.db')
    c = conn.cursor()
    c.execute('''
                INSERT INTO loginDevice (user, id, ip, iat)
                VALUES (?, ?, ?, ?)
            ''', (id, random_id, device_ip, iat))
    conn.commit()
    conn.close()

    # return str(jwt.encode(payload, JWT_SECRET, algorithm='HS256').decode('ascii'))
    JWT_SECRET = read_config()['JWT_SECRET']
    return str(jwt.encode(payload, JWT_SECRET, algorithm='HS256'))

def verify_jwt(token, device_ip):
    try:
        # Find payload into loginDevice table and check device_ip is same or not, also check exp time
        JWT_SECRET = read_config()['JWT_SECRET']
        payload = jwt.decode(token, JWT_SECRET, algorithms=['HS256'])
        conn = sqlite3.connect('main.db')
        c = conn.cursor()
        c.execute('''
                    SELECT * FROM loginDevice WHERE id = ?
                ''', (payload['id'],))
        record = c.fetchone()
        conn.close()

        logs_green(record)
        logs_blue(payload)

        if not record:
            return False
        if record[0] != payload['account']:
            return False
        if record[2] != device_ip:
            return False
        if int(record[3]) != payload['iat']:
            return False
        if payload['exp'] - payload['iat'] != JWT_EXP_TIME:
            return False
        if int(payload['exp']) < int(time.time()):
            return False
        return True
    except Exception as e:
        return False

def generate_jwt_one_day(id, device_ip):
    iat = int(time.time())
    exp = iat + JWT_EXP_TIME_ONE_DAY
    random_id = create_id(32)
    payload = {
        "account": id,
        "iat": iat,
        "exp": exp,
        "id": random_id,
    }

    # Write device info into device database
    conn = sqlite3.connect('main.db')
    c = conn.cursor()
    c.execute('''
                INSERT INTO loginDevice (user, id, ip, iat)
                VALUES (?, ?, ?, ?)
            ''', (id, random_id, device_ip, iat))
    conn.commit()
    conn.close()

    # return str(jwt.encode(payload, JWT_SECRET, algorithm='HS256').decode('ascii'))
    JWT_SECRET = read_config()['JWT_SECRET']
    return str(jwt.encode(payload, JWT_SECRET, algorithm='HS256'))

def verify_jwt_one_day(token, device_ip):
    try:
        # Find payload into loginDevice table and check device_ip is same or not, also check exp time
        JWT_SECRET = read_config()['JWT_SECRET']
        payload = jwt.decode(token, JWT_SECRET, algorithms=['HS256'])
        conn = sqlite3.connect('main.db')
        c = conn.cursor()
        c.execute('''
                    SELECT * FROM loginDevice WHERE id = ?
                ''', (payload['id'],))
        record = c.fetchone()
        conn.close()

        logs_green(record)
        logs_blue(payload)

        if not record:
            return False
        if record[0] != payload['account']:
            return False
        if record[2] != device_ip:
            return False
        if int(record[3]) != payload['iat']:
            return False
        if payload['exp'] - payload['iat'] != JWT_EXP_TIME_ONE_DAY:
            return False
        if int(payload['exp']) < int(time.time()):
            return False
        return True
    except Exception as e:
        return False

def get_user_id_from_token(token):
    JWT_SECRET = read_config()['JWT_SECRET']
    payload = jwt.decode(token, JWT_SECRET, algorithms=['HS256'])
    return payload['account']

def get_user_dorm(std_id):
    conn = sqlite3.connect('main.db')
    c = conn.cursor()
    c.execute('''
                SELECT dorm FROM member WHERE id = ?
            ''', (std_id,))
    record = c.fetchone()
    conn.close()
    dorm = ""
    
    if record:
        dorm = str(record[0])
    else:
        return 'no'
    return dorm

def get_events_reserve_num_by_dorm(events_id, dorm):
    conn = sqlite3.connect('main.db')
    c = conn.cursor()
    c.execute('''
                SELECT * FROM reserve WHERE event_id = ?
            ''', (events_id,))
    reserve_records = c.fetchall()
    conn.close()

    have_reserve_num = 0
    for reserve_record in reserve_records:
        this_reserve_user_dorm = get_user_dorm(reserve_record[1])
        if this_reserve_user_dorm == dorm:
            have_reserve_num += 1
    return have_reserve_num
# print(get_events_reserve_num_by_dorm('9160', 'morn'))

def get_events_list_by_dorm(dorm):
    events_list = []
    events_detail = {
        "event_id": "",
        "event_name": "",
        "date": "",
        "startTime": "",
        "endTime": "",
        "maxReserve": 0,
        "remainReserve": 0,
        "haveReserve": 0,
    }
    conn = sqlite3.connect('main.db')
    c = conn.cursor()
    c.execute('''
                SELECT * FROM events
            ''')
    record = c.fetchall()
    conn.close()

    for event in record:
        events_detail['event_id'] = event[0]
        events_detail['event_name'] = event[1]
        events_detail['date'] = event[2]
        events_detail['startTime'] = event[3]
        events_detail['endTime'] = event[4]
        if dorm == 'sun':
            events_detail['maxReserve'] = event[5]
            events_detail['remainReserve'] = event[5] - get_events_reserve_num_by_dorm(event[0], 'sun')
        elif dorm == 'moon':
            events_detail['maxReserve'] = event[6]
            events_detail['remainReserve'] = event[6] - get_events_reserve_num_by_dorm(event[0], 'moon')
        elif dorm == 'star':
            events_detail['maxReserve'] = event[7]
            events_detail['remainReserve'] = event[7] - get_events_reserve_num_by_dorm(event[0], 'star')
        elif dorm == 'morn':
            events_detail['maxReserve'] = event[8]
            events_detail['remainReserve'] = event[8] - get_events_reserve_num_by_dorm(event[0], 'morn')
        events_detail['haveReserve'] = events_detail['maxReserve'] - events_detail['remainReserve']
        events_list.append(events_detail.copy())
    return events_list
# print(get_events_list_by_dorm('morn'))

def get_user_dorm(id):
    conn = sqlite3.connect('main.db')
    c = conn.cursor()
    c.execute('''
                SELECT dorm FROM member WHERE id = ?
            ''', (id,))
    record = c.fetchone()
    conn.close()
    dorm = ""
    
    if record:
        dorm = str(record[0])
    else:
        return 'no'
    return dorm

def get_user_dorm_and_room(id):
    conn = sqlite3.connect('main.db')
    c = conn.cursor()
    c.execute('''
                SELECT dorm, room, bed FROM member WHERE id = ?
            ''', (id,))
    record = c.fetchone()
    conn.close()
    dorm = ""
    room = ""
    
    if record:
        dorm = str(record[0])
        room = str(record[1])
        bed = str(record[2])
        res = ""
        if dorm == 'sun':
            res = '曉日 / Sun ' + room
        elif dorm == 'moon':
            res = '皓月 / Moon ' + room
        elif dorm == 'star':
            res = '繁星 / Star ' + room
        elif dorm == 'morn':
            res = '辰曦 / Morn ' + room
        return res
    else:
        return 'no'

def get_user_name(id):
    conn = sqlite3.connect('main.db')
    c = conn.cursor()
    c.execute('''
                SELECT name FROM member WHERE id = ?
            ''', (id,))
    record = c.fetchone()
    conn.close()
    name = ""
    
    if record:
        name = str(record[0])
    else:
        return 'no'
    return name

def check_user_if_reserved(id):
    conn = sqlite3.connect('main.db')
    c = conn.cursor()
    c.execute('''
                SELECT * FROM reserve WHERE id = ?
            ''', (id,))
    record = c.fetchone()
    conn.close()

    if record:
        return True
    else:
        return False

def check_user_if_checkin(id):
    conn = sqlite3.connect('main.db')
    c = conn.cursor()
    c.execute('''
                SELECT * FROM checkIn WHERE user = ?
            ''', (id,))
    record = c.fetchone()
    conn.close()

    if record:
        return True
    else:
        return False

def get_user_reserve_record(id):
    conn = sqlite3.connect('main.db')
    c = conn.cursor()
    c.execute('''
                SELECT * FROM reserve WHERE id = ?
            ''', (id,))
    record = c.fetchone()
    conn.close()

    return record[1], record[2], record[3]

def add_reserve_log(id, event_id, parking):
    conn = sqlite3.connect('main.db')
    c = conn.cursor()
    c.execute('''
                INSERT INTO reserve (timestamp, id, event_id, parking)
                VALUES (?, ?, ?, ?)
            ''', (get_nowtime_taipei_time(), id, event_id, parking))
    conn.commit()
    conn.close()

def edit_health_data(id, health):
    is_health = False
    if health['check_1'] == False and health['check_2'] == False:
        is_health = True
    conn = sqlite3.connect('main.db')
    c = conn.cursor()
    c.execute('''
                UPDATE member SET health = ?, details = ? WHERE id = ?
            ''', (is_health, str(health), id))
    conn.commit()
    conn.close()

# Delete reserve log by id
def delete_reserve_log(id):
    conn = sqlite3.connect('main.db')
    c = conn.cursor()
    c.execute('''
                DELETE FROM reserve WHERE id = ?
            ''', (id,))
    conn.commit()
    conn.close()

def healthy_form_check(id):
    conn = sqlite3.connect('main.db')
    c = conn.cursor()
    c.execute('''
                SELECT health FROM member WHERE id = ?
            ''', (id,))
    record = c.fetchone()
    conn.close()

    if record[0] == None:
        return False
    else:
        return True
# print(healthy_form_check('410885045'))

def get_event_name_by_id(id):
    conn = sqlite3.connect('main.db')
    c = conn.cursor()
    c.execute('''
                SELECT event_name FROM events WHERE event_id = ?
            ''', (id,))
    record = c.fetchone()
    conn.close()

    return record[0]

def admin_take_user_data(id):
    # Get data from member table
    conn = sqlite3.connect('main.db')
    c = conn.cursor()
    c.execute('''
                SELECT * FROM member WHERE id = ?
            ''', (id,))
    record = c.fetchone()

    if not record:
        return {
            'error': True
        }
    
    if record[7] == None:
        health_detail = {
            'phone': 'None',
            'check_1': 'None',
            'check_2': 'None'
        }
    else:

        try:
            health_detail = json.loads(str(record[7]).replace("'", '"').replace("False", 'false').replace("True", 'true'))
        except:
            health_detail = {
            'phone': 'Err',
            'check_1': 'Err',
            'check_2': 'Err'
        }
    # print(type(health_detail))

    user_data = {
        'error': False,
        "id": record[0],
        "name": record[1],
        "dorm": record[2],
        "dormO": record[2],
        "room": record[3],
        "health": record[6],
        # "health_detail": health_detail,
        "health_phone": health_detail['phone'],
        "health_rule1": health_detail['check_1'],
        "health_rule2": health_detail['check_2'],
    }

    if user_data['dorm'] == "sun":
        user_data['dorm'] = "曉日"
    elif user_data['dorm'] == "moon":
        user_data['dorm'] = "皓月"
    elif user_data['dorm'] == "star":
        user_data['dorm'] = "繁星"
    elif user_data['dorm'] == "morn":
        user_data['dorm'] = "辰曦"

    # Get data from reserve table
    c.execute('''
                SELECT * FROM reserve WHERE id = ?
            ''', (id,))
    record = c.fetchone()

    if not record:
        user_data['reserved'] = False
    else:
        user_data['reserved'] = True
        user_data['reserve_event'] = get_event_name_by_id(record[2])
        user_data['reserve_parking'] = record[3]
        user_data['reserve_time'] = record[0]

    # Get data from checkIn table
    c.execute('''
                SELECT * FROM checkIn WHERE user = ?
            ''', (id,))
    record = c.fetchone()

    if not record:
        user_data['checkIn'] = False
    else:
        user_data['checkIn'] = True
        user_data['checkIn_time'] = record[0]
        user_data['checkIn_parking'] = record[2]
        user_data['checkIn_bill'] = record[3]
        user_data['checkIn_card'] = record[4]
        user_data['checkIn_visitor_id'] = record[5]
        user_data['checkIn_visitor_phone'] = record[6]
        user_data['checkIn_visitor_start'] = record[7]
        user_data['checkIn_visitor_end'] = record[8]
        user_data['checkIn_visitor_in'] = False
        if user_data['checkIn_visitor_end'] == None and user_data['checkIn_visitor_id'] != None:
            user_data['checkIn_visitor_in'] = True
    return user_data
# print(admin_take_user_data('410873016'))

def admin_edit_user_checkin(id, parking, bill, card):
    # Get user whether check in
    conn = sqlite3.connect('main.db')
    c = conn.cursor()
    c.execute('''
                SELECT * FROM checkIn WHERE user = ?
            ''', (id,))
    record = c.fetchone()

    if not record:
        # Insert a record to checkIn table
        c.execute('''
                    INSERT INTO checkIn (timestamp, user, parking, bill, card)
                    VALUES (?, ?, ?, ?, ?)
                ''', (get_nowtime_taipei_time(), id, parking, bill, card))
        conn.commit()
        conn.close()
    else:
        # Update a record to checkIn table
        c.execute('''
                    UPDATE checkIn SET parking = ?, bill = ?, card = ? WHERE user = ?
                ''', (parking, bill, card, id))
        conn.commit()
        conn.close()
    return True

def admin_edit_user_visitor(id, visitor_id, visitor_phone):
    # Get user whether check in
    conn = sqlite3.connect('main.db')
    c = conn.cursor()
    c.execute('''
                SELECT * FROM checkIn WHERE user = ?
            ''', (id,))
    record = c.fetchone()

    if not record:
        return False
    else:
        if record[5] != None:
            return True
        # Update a record to checkIn table
        c.execute('''
                    UPDATE checkIn SET visitor_id = ?, visitor_phone = ?, visitor_start = ? WHERE user = ?
                ''', (visitor_id, visitor_phone, get_nowtime_taipei_time(), id))
        conn.commit()
        conn.close()
        return True

def admin_visitor_checkout(id):
    # Get user whether check in
    conn = sqlite3.connect('main.db')
    c = conn.cursor()
    c.execute('''
                SELECT * FROM checkIn WHERE visitor_phone = ?
            ''', (id,))
    record = c.fetchone()

    if not record:
        return False, None
    else:
        user = record[1]
        # Update a record to checkIn table
        c.execute('''
                    UPDATE checkIn SET visitor_end = ? WHERE visitor_phone = ?
                ''', (get_nowtime_taipei_time(), id))
        conn.commit()
        conn.close()
        return True, user

def admin_visitor_checkout_by_stuid(id):
    # Get user whether check in
    conn = sqlite3.connect('main.db')
    c = conn.cursor()
    c.execute('''
                SELECT * FROM checkIn WHERE user = ?
            ''', (id,))
    record = c.fetchone()

    if not record:
        return False, None
    else:
        user = record[1]
        # Update a record to checkIn table
        c.execute('''
                    UPDATE checkIn SET visitor_end = ? WHERE user = ?
                ''', (get_nowtime_taipei_time(), id))
        conn.commit()
        conn.close()
        return True, user

def admin_get_reserve_status():
    conn = sqlite3.connect('main.db')
    c = conn.cursor()
    c.execute('''
                SELECT e.event_name, m.id, m.dorm, r.parking, m.health, c.timestamp, c.visitor_end, c.visitor_start
                FROM reserve AS r
                LEFT JOIN member AS m ON m.id=r.id
                LEFT JOIN events AS e ON r.event_id=e.event_id
                LEFT JOIN checkIn AS c ON m.id=c.user
                ORDER BY e.event_name ASC
            ''')
    record = c.fetchall()
    conn.close()

    res = []
    for row in record:
        tmp = {
            "d": row[0],
            "s": row[1],
            "b": row[2],
            "p": row[3],
            "h": row[4],
            "c": row[5],
            "vi": 'n',
            "vs": row[7],
            "de": row[1],
        }
        if tmp['p'] == 'no':
            tmp['p'] = 'n'
        if tmp['c'] == None:
            tmp['c'] = 'n'
        if tmp['vs'] != None and tmp['vs'] == None:
            tmp['vi'] = 'y'
        if tmp['de'] != None:
            res.append(tmp)
    return res

def admin_delete_reserve(id):
    # Check whether the user is reserved
    conn = sqlite3.connect('main.db')
    c = conn.cursor()
    c.execute('''
                SELECT * FROM reserve WHERE id = ?
            ''', (id,))
    record = c.fetchone()

    if not record:
        return False

    c.execute('''
                DELETE FROM reserve WHERE id = ?
            ''', (id,))
    conn.commit()
    conn.close()
    return True

def admin_download_csv():
    import openpyxl
    conn = sqlite3.connect('main.db')
    c = conn.cursor()
    c.execute('''
                SELECT m.id, m.name, m.dorm, m.room, m.health, m.details, r.timestamp, e.event_name, r.parking, c.timestamp, c.parking, c.bill, c.card, c.visitor_id, c.visitor_phone, c.visitor_start, c.visitor_end
                FROM member AS m
                LEFT JOIN reserve AS r ON m.id=r.id
                LEFT JOIN events AS e ON r.event_id=e.event_id
                LEFT JOIN checkIn AS c ON m.id=c.user
            ''')
    record = c.fetchall()
    conn.close()

    # Check if a folder exists named "csv_output"
    if not os.path.exists("csv_output"):
        os.makedirs("csv_output")
    
    # Write data to csv file
    fileName = "C:\\Users\\littl\\Program\\NTPU_Dorm_Reserve_Backend\\csv_output\\"
    fileName += get_nowtime_taipei_time().replace(" ", "").replace("-", "").replace(":", "") + ".csv"
    with open(fileName, 'w', newline='', encoding="utf-8") as csvfile:
        fieldnames = ['學號', '姓名', '宿舍', '寢室床號', '健康狀態', '聲明書內容', '預約時間', '預約事件', '預約停車券車牌', '報到時間', '領取停車券', '繳費', '臨時卡', '訪客身分證', '訪客電話', '訪客開始時間', '訪客結束時間']
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        for row in record:
            writer.writerow({'學號': row[0], '姓名': row[1], '宿舍': row[2], '寢室床號': row[3], '健康狀態': row[4], '聲明書內容': row[5], '預約時間': row[6], '預約事件': row[7], '預約停車券車牌': row[8], '報到時間': row[9], '領取停車券': row[10], '繳費': row[11], '臨時卡': row[12], '訪客身分證': row[13], '訪客電話': row[14], '訪客開始時間': row[15], '訪客結束時間': row[16]})

    xlName = "C:\\Users\\littl\\Program\\NTPU_Dorm_Reserve_Backend\\csv_output\\"
    xlName += get_nowtime_taipei_time().replace(" ", "").replace("-", "").replace(":", "") + ".xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active

    with open(fileName, 'r', newline='', encoding="utf-8") as f:
        reader = csv.reader(f, delimiter=',')
        for row in reader:
            ws.append(row)

    ws.freeze_panes = ws['B2']
    wb.save(xlName)

    # wb = openpyxl.Workbook()
    # ws = wb.active
    # with open(fileName, 'w', newline='', encoding="utf-8") as f:
    #     reader = csv.reader(f, delimiter=',')
    # for row in reader:
    #     ws.append(row)
    # wb.save(xlName)

    return xlName

def admin_add_user(id, name, dorm, room):
    # Check this user in member table
    conn = sqlite3.connect('main.db')
    c = conn.cursor()
    c.execute('''
                SELECT * FROM member WHERE id = ?
            ''', (id,))
    record = c.fetchone()
    if record:
        return False
    
    # Insert a record to member table
    c.execute('''
                INSERT INTO member (id, name, dorm, room, status)
                VALUES (?, ?, ?, ?, 'normal')
            ''', (id, name, dorm, room))
    # c.execute('''
    #             INSERT INTO accounts (id, passwd, status)
    #             VALUES (?, ?, 'normal')
    #         ''', (id, 'testtesttest'))
    conn.commit()
    conn.close()
    return True

def admin_edit_user(id, name, dorm, room):
    # Check this user in member table
    conn = sqlite3.connect('main.db')
    c = conn.cursor()
    c.execute('''
                SELECT * FROM member WHERE id = ?
            ''', (id,))
    record = c.fetchone()
    print(record)
    if not record:
        return False

    # Edit a record to member table
    c.execute('''
                UPDATE member SET name = ?, dorm = ?, room = ? WHERE id = ?
            ''', (name, dorm, room, id))
    conn.commit()
    conn.close()
    return True

def admin_delete_user(id):
    print('Delete work in progress, ' + id)
    # Check this user in member table
    conn = sqlite3.connect('main.db')
    c = conn.cursor()
    c.execute('''
                SELECT * FROM reserve WHERE id = ?
            ''', (id,))
    record = c.fetchone()
    if not record:
        print("No such reserve")
    else:
        # Delete a record from member table
        c.execute('''
                    DELETE FROM reserve WHERE id = ?
                ''', (id,))
        print('Delete reserve record')

    # Check this user in member table
    c.execute('''
                SELECT * FROM member WHERE id = ?
            ''', (id,))
    record = c.fetchone()
    if not record:
        print("No such member")
        return False
    
    # Delete a record from member table
    c.execute('''
                DELETE FROM member WHERE id = ?
            ''', (id,))
    conn.commit()
    conn.close()
    return True

def admin_delete_user_checkin(id):
    print('Delete work in progress, ' + id)
    # Check this user in member table
    conn = sqlite3.connect('main.db')
    c = conn.cursor()
    c.execute('''
                SELECT * FROM checkIn WHERE user = ?
            ''', (id,))
    record = c.fetchone()
    if not record:
        print("No such checkIn")
    else:
        # Delete a record from member table
        c.execute('''
                    DELETE FROM checkIn WHERE user = ?
                ''', (id,))
        print('Delete checkIn record')
    conn.commit()
    conn.close()
    return True

def get_parking_couple():
    conn = sqlite3.connect('main.db')
    c = conn.cursor()
    c.execute('''
                SELECT e.event_name, r.parking
                FROM reserve AS r
                LEFT JOIN events AS e ON r.event_id=e.event_id
            ''')
    record = c.fetchall()
    conn.close()

    counter = {}
    for i in record:
        if i[0] not in counter and i[1] != 'no':
            counter[i[0]] = 1
        elif i[1] != 'no':
            counter[i[0]] += 1

    return json.loads(json.dumps(counter, sort_keys=True))

def get_visitor_inside():
    # Get visitor_start is not None in CheckIn table
    conn = sqlite3.connect('main.db')
    c = conn.cursor()
    c.execute('''
                SELECT user, visitor_id, visitor_phone, visitor_start, visitor_end FROM checkIn WHERE visitor_start IS NOT NULL
            ''')
    record = c.fetchall()
    conn.close()

    res = []

    for i in record:
        if i[4] == None:
            res.append({
                'user': i[0],
                'dorm': get_user_dorm_and_room(i[0]),
                'name': get_user_name(i[0]),
                'visitor_id': i[1],
                'visitor_phone': i[2],
                'visitor_start': i[3],
            })
    return res

def get_checkin_status():
    conn = sqlite3.connect('main.db')
    c = conn.cursor()
    c.execute('''
                SELECT c.user, e.event_name, m.dorm
                FROM checkIn AS c
                LEFT JOIN member AS m ON m.id=c.user
                LEFT JOIN reserve AS r ON r.id=m.id
                LEFT JOIN events AS e ON e.event_id=r.event_id
            ''')
    record = c.fetchall()
    conn.close()

    counter = {
        'all': [],
        'sun': [],
        'moon': [],
        'star': [],
        'morn': []
    }

    tmp = {}

    for i in record:
        if i[1] not in tmp:
            if i[1] == None:
                if 'other' not in tmp:
                    tmp['other'] = 1
                else:
                    tmp['other'] += 1
            else:
                tmp[i[1]] = 1
        else:
            if i[1] == None:
                tmp['other'] += 1
            else:
                tmp[i[1]] += 1

    counter['all'] = json.loads(json.dumps(tmp, sort_keys=True))

    dormList = ['sun', 'moon', 'star', 'morn']
    for d in dormList:
        tmp = {}
        for i in record:
            if i[2] == d:
                if i[1] not in tmp:
                    if i[1] == None:
                        if 'other' not in tmp:
                            tmp['other'] = 1
                        else:
                            tmp['other'] += 1
                    else:
                        tmp[i[1]] = 1
                else:
                    if i[1] == None:
                        tmp['other'] += 1
                    else:
                        tmp[i[1]] += 1
        counter[d] = json.loads(json.dumps(tmp, sort_keys=True))

    return counter

def base64_encode(text):
    return base64.b64encode(text.encode('utf-8')).decode('utf-8')

def base64_decode(text):
    return base64.b64decode(text).decode('utf-8')

def check_is_num(input):
    for i in range(len(input)):
        if ord(input[i]) >= 48 and ord(input[i]) <= 57:
            continue
        else:
            return False
    return True

def logs_blue(msg):
    print('\033[94m' + str(msg) + '\033[0m')

def logs_red(msg):
    print('\033[91m' + str(msg) + '\033[0m')

def logs_green(msg):
    print('\033[92m' + str(msg) + '\033[0m')

def error_403(mes, session, remote_addr, url, events_id):
    json_data = {
        "code": 403,
        "status": "Forbidden",
        "message": mes,
        "session": session,
    }
    write_log(remote_addr, url, events_id, '', str(json_data))
    logs_red('error_403')
    return json_data

def ok_200(mes, session, remote_addr, url, events_id):
    json_data = {
        "code": 200,
        "status": "OK",
        "message": mes,
        "session": session,
    }
    write_log(remote_addr, url, events_id, '', str(json_data))
    logs_green('ok_200')
    return json_data

def error_500(mes, session, remote_addr, url, events_id):
    json_data = {
        "code": 500,
        "status": "Internal Server Error",
        "message": mes,
        "session": session,
    }
    write_log(remote_addr, url, events_id, '', str(json_data))
    logs_red('error_500')
    return json_data

def read_config():
    with open('config.json', 'r', encoding='utf-8') as f:
        config = json.load(f)
    return config
